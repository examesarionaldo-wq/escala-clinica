import calendar
import io
import json
from dataclasses import dataclass
from datetime import date
from typing import Dict, List, Tuple

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from ortools.sat.python import cp_model

st.set_page_config(page_title="Escala da Clínica", page_icon="🐾", layout="wide")

MESES = {
    1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
    5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
    9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro",
}
DIAS_SEMANA = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]

DEFAULT_STAFF = ["Ana", "Danielle", "Suzana", "Nicolle", "Plantonista extra"]


def init_state() -> None:
    defaults = {
        "staff": DEFAULT_STAFF.copy(),
        "schedule": None,
        "warnings": [],
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


@dataclass
class Config:
    year: int
    month: int
    holidays: List[int]
    fixed_nights: Dict[int, str]
    unavailable: Dict[str, List[int]]
    first_sunday: str


def month_days(year: int, month: int) -> List[int]:
    return list(range(1, calendar.monthrange(year, month)[1] + 1))


def weekday(year: int, month: int, day: int) -> int:
    return date(year, month, day).weekday()  # 0=segunda, 6=domingo


def build_solver(config: Config) -> Tuple[cp_model.CpModel, Dict, List[int], List[str], List[str]]:
    model = cp_model.CpModel()
    days = month_days(config.year, config.month)
    staff = st.session_state.staff
    shifts = ["Dia 07h–16h", "Dia 10h–19h", "Dom/Feriado 07h–19h", "Noite 19h–07h"]
    x = {}

    for p in staff:
        for d in days:
            for s in shifts:
                x[p, d, s] = model.NewBoolVar(f"x_{p}_{d}_{s}")

    ana = "Ana"
    dani = "Danielle"
    suzana = "Suzana"
    nicolle = "Nicolle"
    extra = "Plantonista extra"

    for d in days:
        wd = weekday(config.year, config.month, d)
        special = wd == 6 or d in config.holidays

        if special:
            model.Add(sum(x[p, d, "Dom/Feriado 07h–19h"] for p in [ana, dani, suzana]) == 1)
            for p in staff:
                model.Add(x[p, d, "Dia 07h–16h"] == 0)
                model.Add(x[p, d, "Dia 10h–19h"] == 0)
            model.Add(x[nicolle, d, "Dom/Feriado 07h–19h"] == 0)
            model.Add(x[extra, d, "Dom/Feriado 07h–19h"] == 0)
        elif wd == 5:  # sábado
            model.Add(x[ana, d, "Dia 07h–16h"] + x[dani, d, "Dia 07h–16h"] == 1)
            model.Add(x[ana, d, "Dia 10h–19h"] + x[dani, d, "Dia 10h–19h"] == 1)
            model.Add(x[ana, d, "Dia 07h–16h"] + x[ana, d, "Dia 10h–19h"] == 1)
            model.Add(x[dani, d, "Dia 07h–16h"] + x[dani, d, "Dia 10h–19h"] == 1)
            for p in [suzana, nicolle, extra]:
                model.Add(x[p, d, "Dia 07h–16h"] == 0)
                model.Add(x[p, d, "Dia 10h–19h"] == 0)
            for p in staff:
                model.Add(x[p, d, "Dom/Feriado 07h–19h"] == 0)
        else:  # segunda a sexta
            model.Add(x[ana, d, "Dia 07h–16h"] + x[dani, d, "Dia 07h–16h"] == 1)
            model.Add(x[suzana, d, "Dia 10h–19h"] == 1)
            for p in [ana, dani, nicolle, extra]:
                model.Add(x[p, d, "Dia 10h–19h"] == 0)
            for p in [suzana, nicolle, extra]:
                model.Add(x[p, d, "Dia 07h–16h"] == 0)
            for p in staff:
                model.Add(x[p, d, "Dom/Feriado 07h–19h"] == 0)

        # Uma pessoa não pode assumir dois turnos diurnos no mesmo dia.
        for p in staff:
            model.Add(
                x[p, d, "Dia 07h–16h"] + x[p, d, "Dia 10h–19h"] + x[p, d, "Dom/Feriado 07h–19h"] <= 1
            )

        # Exatamente uma pessoa à noite.
        model.Add(sum(x[p, d, "Noite 19h–07h"] for p in staff) == 1)

        # Nicolle apenas nas noites fixadas; plantonista extra apenas fim de semana.
        if d not in config.fixed_nights or config.fixed_nights[d] != nicolle:
            model.Add(x[nicolle, d, "Noite 19h–07h"] == 0)
        if wd not in (5, 6):
            model.Add(x[extra, d, "Noite 19h–07h"] == 0)
        model.Add(x[suzana, d, "Noite 19h–07h"] == 0)

    # Fixar noites informadas.
    for d, person in config.fixed_nights.items():
        if d in days and person in staff:
            model.Add(x[person, d, "Noite 19h–07h"] == 1)

    # Indisponibilidades.
    for person, blocked_days in config.unavailable.items():
        if person not in staff:
            continue
        for d in blocked_days:
            if d in days:
                for s in shifts:
                    model.Add(x[person, d, s] == 0)

    # Primeiro domingo.
    sundays = [d for d in days if weekday(config.year, config.month, d) == 6]
    if sundays and config.first_sunday in [ana, dani, suzana]:
        model.Add(x[config.first_sunday, sundays[0], "Dom/Feriado 07h–19h"] == 1)

    # Regra obrigatória: a mesma veterinária não pode fazer dois domingos consecutivos.
    for p in [ana, dani, suzana]:
        for i in range(len(sundays) - 1):
            model.Add(
                x[p, sundays[i], "Dom/Feriado 07h–19h"]
                + x[p, sundays[i + 1], "Dom/Feriado 07h–19h"]
                <= 1
            )

    penalties = []

    # Equilibrar os plantões noturnos de Ana e Danielle.
    # A diferença é minimizada com prioridade alta: igualdade quando possível
    # e diferença mínima quando o número de noites disponíveis for ímpar.
    nights_ana = sum(x[ana, d, "Noite 19h–07h"] for d in days)
    nights_dani = sum(x[dani, d, "Noite 19h–07h"] for d in days)
    diff_nights = model.NewIntVar(0, len(days), "diff_nights")
    model.AddAbsEquality(diff_nights, nights_ana - nights_dani)
    penalties.append(diff_nights * 1000)

    # Equilibrar todos os plantões diurnos de Ana e Danielle, incluindo
    # turnos de segunda a sábado, domingos e feriados.
    days_ana = sum(
        x[ana, d, "Dia 07h–16h"]
        + x[ana, d, "Dia 10h–19h"]
        + x[ana, d, "Dom/Feriado 07h–19h"]
        for d in days
    )
    days_dani = sum(
        x[dani, d, "Dia 07h–16h"]
        + x[dani, d, "Dia 10h–19h"]
        + x[dani, d, "Dom/Feriado 07h–19h"]
        for d in days
    )
    diff_days = model.NewIntVar(0, len(days), "diff_days")
    model.AddAbsEquality(diff_days, days_ana - days_dani)
    penalties.append(diff_days * 1000)

    # Distribuir melhor os plantões ao longo do mês.
    # Além do equilíbrio mensal, o solver tenta manter Ana e Danielle
    # próximas também dentro de cada semana (segunda a domingo). Isso evita
    # concentrar uma pessoa em vários diurnos numa semana e compensar com
    # vários noturnos na semana seguinte.
    week_groups = {}
    for d in days:
        dt = date(config.year, config.month, d)
        monday = dt.toordinal() - dt.weekday()
        week_groups.setdefault(monday, []).append(d)

    for week_index, week_days in enumerate(week_groups.values(), start=1):
        week_day_ana = sum(
            x[ana, d, "Dia 07h–16h"]
            + x[ana, d, "Dia 10h–19h"]
            + x[ana, d, "Dom/Feriado 07h–19h"]
            for d in week_days
        )
        week_day_dani = sum(
            x[dani, d, "Dia 07h–16h"]
            + x[dani, d, "Dia 10h–19h"]
            + x[dani, d, "Dom/Feriado 07h–19h"]
            for d in week_days
        )
        week_night_ana = sum(x[ana, d, "Noite 19h–07h"] for d in week_days)
        week_night_dani = sum(x[dani, d, "Noite 19h–07h"] for d in week_days)

        max_week = len(week_days)
        diff_week_days = model.NewIntVar(0, max_week, f"diff_week_days_{week_index}")
        diff_week_nights = model.NewIntVar(0, max_week, f"diff_week_nights_{week_index}")
        model.AddAbsEquality(diff_week_days, week_day_ana - week_day_dani)
        model.AddAbsEquality(diff_week_nights, week_night_ana - week_night_dani)
        penalties.append(diff_week_days * 120)
        penalties.append(diff_week_nights * 120)

    # Penalizar sequências longas do mesmo tipo de plantão para Ana e Dani.
    # Janelas de 4 dias com 3 ou 4 diurnos (ou noturnos) recebem penalidade
    # crescente. Não é uma proibição rígida, para não inviabilizar a escala
    # quando houver indisponibilidades ou noites fixas.
    for p in [ana, dani]:
        day_work = {}
        for d in days:
            day_work[d] = (
                x[p, d, "Dia 07h–16h"]
                + x[p, d, "Dia 10h–19h"]
                + x[p, d, "Dom/Feriado 07h–19h"]
            )

        for start in range(1, max(days) - 2):
            window = [d for d in range(start, start + 4) if d in days]
            if len(window) != 4:
                continue

            excess_days = model.NewIntVar(0, 2, f"excess_days_{p}_{start}")
            excess_nights = model.NewIntVar(0, 2, f"excess_nights_{p}_{start}")
            model.Add(excess_days >= sum(day_work[d] for d in window) - 2)
            model.Add(excess_nights >= sum(x[p, d, "Noite 19h–07h"] for d in window) - 2)
            penalties.append(excess_days * 180)
            penalties.append(excess_nights * 180)

    # Equilibrar domingos/feriados.
    specials = [d for d in days if weekday(config.year, config.month, d) == 6 or d in config.holidays]
    counts = {}
    for p in [ana, dani, suzana]:
        counts[p] = sum(x[p, d, "Dom/Feriado 07h–19h"] for d in specials)
    for p1, p2 in [(ana, dani), (ana, suzana), (dani, suzana)]:
        diff = model.NewIntVar(0, len(specials), f"diff_special_{p1}_{p2}")
        model.AddAbsEquality(diff, counts[p1] - counts[p2])
        penalties.append(diff * 8)

    # Minimizar plantonista extra e dobradinhas.
    # Também equilibrar a quantidade de dobradinhas entre Ana e Danielle.
    double_vars = {ana: [], dani: []}
    # Dobradinhas recebem penalidade muito alta para que o solver só as use
    # quando forem realmente necessárias para manter a escala viável. A
    # distribuição semanal nunca deve ser "melhorada" criando dobradinhas.
    # Regra obrigatória da dobradinha:
    # - é proibido trabalhar durante o dia e iniciar a noite às 19h no mesmo dia;
    # - a única dobradinha permitida começa às 19h e termina às 16h do dia seguinte,
    #   isto é: Noite 19h–07h + Dia 07h–16h do dia seguinte.
    for d in days:
        penalties.append(x[extra, d, "Noite 19h–07h"] * 10000)

        for p in [ana, dani]:
            night = x[p, d, "Noite 19h–07h"]

            # Proíbe a sequência incorreta 07h/10h durante o dia + noite às 19h.
            # Assim, ninguém pode começar às 07h e seguir até às 16h do outro dia.
            model.Add(night + x[p, d, "Dia 07h–16h"] <= 1)
            model.Add(night + x[p, d, "Dia 10h–19h"] <= 1)
            model.Add(night + x[p, d, "Dom/Feriado 07h–19h"] <= 1)

            if d < days[-1]:
                next_day = d + 1
                early_next = x[p, next_day, "Dia 07h–16h"]
                late_next = x[p, next_day, "Dia 10h–19h"]
                special_next = x[p, next_day, "Dom/Feriado 07h–19h"]

                # Após a noite, a única continuidade possível seria 07h–16h.
                # Porém, essa dobradinha é permitida APENAS de sexta para sábado.
                model.Add(night + late_next <= 1)
                model.Add(night + special_next <= 1)

                is_friday_night = weekday(config.year, config.month, d) == 4 and weekday(config.year, config.month, next_day) == 5
                if is_friday_night:
                    double = model.NewBoolVar(f"double_{p}_{d}")
                    model.Add(double <= night)
                    model.Add(double <= early_next)
                    model.Add(double >= night + early_next - 1)
                    double_vars[p].append(double)
                    penalties.append(double * 5000)

                    # Se Ana ou Danielle fizer sexta à noite + sábado 07h–16h,
                    # deve folgar o domingo inteiro seguinte.
                    if d + 2 <= days[-1]:
                        sunday = d + 2
                        if weekday(config.year, config.month, sunday) == 6:
                            model.Add(x[p, sunday, "Dom/Feriado 07h–19h"] + double <= 1)
                            model.Add(x[p, sunday, "Noite 19h–07h"] + double <= 1)
                else:
                    # Em qualquer outro dia da semana, a dobradinha é proibida.
                    model.Add(night + early_next <= 1)

    # Equilíbrio das dobradinhas entre Ana e Danielle.
    # O solver busca igualdade exata; se o total necessário for ímpar,
    # minimiza a diferença para no máximo 1 sempre que a escala permitir.
    doubles_ana = sum(double_vars[ana]) if double_vars[ana] else 0
    doubles_dani = sum(double_vars[dani]) if double_vars[dani] else 0
    max_doubles = max(len(double_vars[ana]), len(double_vars[dani]), 1)
    diff_doubles = model.NewIntVar(0, max_doubles, "diff_doubles")
    model.AddAbsEquality(diff_doubles, doubles_ana - doubles_dani)
    penalties.append(diff_doubles * 4000)

    model.Minimize(sum(penalties))
    return model, x, days, staff, shifts


def solve_schedule(config: Config):
    model, x, days, staff, shifts = build_solver(config)
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 20
    solver.parameters.num_search_workers = 8
    status = solver.Solve(model)

    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        return None, ["Não foi possível gerar uma escala com todas as regras e indisponibilidades informadas."]

    rows = []
    for d in days:
        wd = DIAS_SEMANA[weekday(config.year, config.month, d)]
        row = {"Dia": d, "Semana": wd}
        for s in shifts:
            assigned = [p for p in staff if solver.Value(x[p, d, s]) == 1]
            row[s] = assigned[0] if assigned else "—"
        rows.append(row)

    df = pd.DataFrame(rows)
    return df, []


def to_excel(df: pd.DataFrame, config: Config) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f"{MESES[config.month]} {config.year}"

    headers = ["Turno"] + [f"{d}\n{DIAS_SEMANA[weekday(config.year, config.month, d)]}" for d in month_days(config.year, config.month)]
    for col, value in enumerate(headers, 1):
        cell = ws.cell(1, col, value)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    shift_rows = [
        ("07h–16h", "Dia 07h–16h"),
        ("10h–19h", "Dia 10h–19h"),
        ("Dom/Feriado 07h–19h", "Dom/Feriado 07h–19h"),
        ("Noite 19h–07h", "Noite 19h–07h"),
    ]
    for row_idx, (label, key) in enumerate(shift_rows, 2):
        ws.cell(row_idx, 1, label).font = Font(bold=True)
        for col_idx, day in enumerate(month_days(config.year, config.month), 2):
            value = df.loc[df["Dia"] == day, key].iloc[0]
            cell = ws.cell(row_idx, col_idx, value)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 24
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = 13
    for row in range(1, 6):
        ws.row_dimensions[row].height = 34

    summary = wb.create_sheet("Resumo")
    summary.append(["Veterinária", "Diurnos", "Noturnos", "Domingos/Feriados"])
    for cell in summary[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for person in st.session_state.staff:
        diurnos = int((df["Dia 07h–16h"] == person).sum() + (df["Dia 10h–19h"] == person).sum())
        noturnos = int((df["Noite 19h–07h"] == person).sum())
        especiais = int((df["Dom/Feriado 07h–19h"] == person).sum())
        summary.append([person, diurnos, noturnos, especiais])

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def config_to_json(config: Config) -> bytes:
    payload = {
        "year": config.year,
        "month": config.month,
        "holidays": config.holidays,
        "fixed_nights": config.fixed_nights,
        "unavailable": config.unavailable,
        "first_sunday": config.first_sunday,
        "staff": st.session_state.staff,
    }
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")


init_state()
st.title("🐾 Gerador de Escalas da Clínica")
st.caption("Geração automática, conferência e exportação para Excel.")

with st.sidebar:
    st.header("Configuração")
    year = st.number_input("Ano", min_value=2025, max_value=2035, value=2026, step=1)
    month = st.selectbox("Mês", list(MESES.keys()), index=7, format_func=lambda m: MESES[m])
    days = month_days(int(year), int(month))

    holidays = st.multiselect("Feriados do mês", days, help="Selecione apenas os números dos dias.")
    first_sunday = st.selectbox("Primeiro domingo", ["Ana", "Danielle", "Suzana"])

    st.subheader("Noites fixas")
    fixed_night_days = st.multiselect("Dias de Nicolle", days, default=[d for d in [17, 24, 25, 27, 31] if d in days])
    fixed_nights = {int(d): "Nicolle" for d in fixed_night_days}

    st.subheader("Indisponibilidades")
    unavailable = {}
    for person in ["Ana", "Danielle", "Suzana", "Nicolle"]:
        unavailable[person] = st.multiselect(f"Folgas de {person}", days, key=f"off_{person}_{year}_{month}")

    generate = st.button("Gerar escala", type="primary", use_container_width=True)

config = Config(
    year=int(year),
    month=int(month),
    holidays=[int(d) for d in holidays],
    fixed_nights=fixed_nights,
    unavailable={p: [int(d) for d in ds] for p, ds in unavailable.items()},
    first_sunday=first_sunday,
)

if generate:
    df, warnings = solve_schedule(config)
    st.session_state.schedule = df
    st.session_state.warnings = warnings

if st.session_state.warnings:
    for message in st.session_state.warnings:
        st.error(message)

if st.session_state.schedule is not None:
    df = st.session_state.schedule.copy()
    st.success("Escala gerada com sucesso.")

    tabs = st.tabs(["Escala", "Resumo", "Exportar"])
    with tabs[0]:
        visible = df.copy()
        st.dataframe(visible, use_container_width=True, hide_index=True)
        st.info("A edição manual direta será adicionada na próxima versão. Nesta versão, altere folgas e noites fixas e gere novamente.")

    with tabs[1]:
        summary_rows = []
        for person in st.session_state.staff:
            summary_rows.append({
                "Veterinária": person,
                "Diurnos": int((df["Dia 07h–16h"] == person).sum() + (df["Dia 10h–19h"] == person).sum()),
                "Noturnos": int((df["Noite 19h–07h"] == person).sum()),
                "Domingos/Feriados": int((df["Dom/Feriado 07h–19h"] == person).sum()),
            })
        st.dataframe(pd.DataFrame(summary_rows), use_container_width=True, hide_index=True)

    with tabs[2]:
        excel_bytes = to_excel(df, config)
        st.download_button(
            "Baixar escala em Excel",
            data=excel_bytes,
            file_name=f"escala_{config.year}_{config.month:02d}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        st.download_button(
            "Salvar configurações em JSON",
            data=config_to_json(config),
            file_name=f"configuracao_{config.year}_{config.month:02d}.json",
            mime="application/json",
            use_container_width=True,
        )
else:
    st.info("Defina o mês, as folgas e as noites fixas na barra lateral e clique em **Gerar escala**.")

with st.expander("Regras aplicadas nesta versão"):
    st.markdown(
        """
- Segunda a sexta: Suzana das 10h às 19h; Ana ou Danielle das 7h às 16h.
- Sábado: Ana e Danielle, uma em cada horário diurno.
- Domingo e feriado: uma veterinária entre Ana, Danielle e Suzana.
- Nicolle trabalha apenas nas noites fixadas.
- Demais noites divididas entre Ana e Danielle.
- Plantonista extra apenas em noites de sábado ou domingo e somente quando necessário.
- Ana e Danielle devem ter o mesmo número de plantões diurnos e noturnos; quando o empate exato não for possível, o sistema usa a menor diferença possível.
- O sistema minimiza plantonista extra e dobradinhas.
- A única dobradinha permitida é de sexta à noite para sábado das 07h às 16h.
- Ana e Danielle devem ter o mesmo número de dobradinhas; quando o total necessário for ímpar, o sistema mantém a diferença mínima possível.
- A mesma veterinária não pode trabalhar em dois domingos consecutivos.
"""
    )
